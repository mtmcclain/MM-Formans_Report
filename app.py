import streamlit as st
import openpyxl
import pandas as pd
import tempfile
import os
from io import BytesIO
from datetime import date, datetime

# ─── Load the employee master list ──────────────────────────────────────────
try:
    df_employees = pd.read_excel("EmployeeNames.xlsx")
    employee_list = df_employees["Employee Name"].tolist()
    craft_dict = dict(zip(df_employees["Employee Name"], df_employees["Craft"]))
except Exception as e:
    st.error(f"Could not load EmployeeNames.xlsx!\nError: {e}")
    st.stop()

st.title("Pipefitting Contractor Report Automation")

tab1, tab2 = st.tabs(["Daily Foreman's Report", "Weekly Timesheet"])

# ─── DAILY FOREMAN'S REPORT ─────────────────────────────────────────────────
with tab1:
    # Static company header to match Excel
    st.markdown("""
    **MARTIN MECHANICAL CORPORATION**  
    1419 Junction Ave, Schererville, IN 46375  
    Phone: 219-322-7333 / Fax: 219-322-7337  
    martinmech@martin-mech.com  
    """)

    st.header("FOREMAN'S DAILY REPORT")

    st.subheader("Work Location")

    # Use radio for single selection (cleaner than forcing checkboxes)
    work_state = st.radio(
        "Select the state where the work was performed:",
        options=["Indiana", "Illinois"],
        index=0,  # 0 = Indiana by default
        horizontal=True,  # side by side
        key="work_state_radio"
    )    
        
    # Date and Day
    report_date = st.date_input("DATE:", value=date.today())
    day_name = datetime.combine(report_date, datetime.min.time()).strftime("%A")
    st.text_input("DAY:", value=day_name, disabled=True)

    # Job details
    job_name = st.text_input("JOB NAME:")
    job_number = st.text_input("JOB NUMBER:")
    description = st.text_input("DESCRIPTION:")

    # Employees (your existing dynamic section - unchanged)
    st.subheader("Employees")
    if 'employees' not in st.session_state:
        st.session_state.employees = []
    if 'add_counter' not in st.session_state:
        st.session_state.add_counter = 0

    for i in range(len(st.session_state.employees)):
        cols = st.columns([3, 2, 1, 1, 1, 0.8])
        with cols[0]:
            selected_name = st.selectbox(
                "Name",
                options=[""] + employee_list,
                index=employee_list.index(st.session_state.employees[i]["name"]) + 1 
                      if st.session_state.employees[i]["name"] in employee_list else 0,
                key=f"emp_name_{i}"
            )
        if selected_name and selected_name != "":
            craft = craft_dict.get(selected_name, "")
            with cols[1]:
                st.text_input("Craft", value=craft, disabled=True, key=f"craft_{i}")
            with cols[2]:
                st_h = st.number_input("ST", min_value=0.0, step=0.5,
                                      value=st.session_state.employees[i]["st"], key=f"st_{i}")
            with cols[3]:
                one5_h = st.number_input("x1.5 OT", min_value=0.0, step=0.5,
                                        value=st.session_state.employees[i]["one5"], key=f"one5_{i}")
            with cols[4]:
                dt_h = st.number_input("x2 OT", min_value=0.0, step=0.5,
                                      value=st.session_state.employees[i]["dt"], key=f"dt_{i}")
            with cols[5]:
                if st.button("🗑", key=f"remove_{i}", help="Remove this employee"):
                    del st.session_state.employees[i]
                    st.rerun()
            st.session_state.employees[i] = {
                "name": selected_name,
                "craft": craft,
                "st": st_h,
                "one5": one5_h,
                "dt": dt_h
            }
        else:
            del st.session_state.employees[i]
            st.rerun()

    st.markdown("**Add next employee:**")
    cols = st.columns([3, 2, 1, 1, 1, 0.8])
    add_suffix = f"_{st.session_state.add_counter}"
    with cols[0]:
        new_name = st.selectbox("Name", options=[""] + employee_list, key=f"new_emp_name{add_suffix}")
    if new_name and new_name != "":
        new_craft = craft_dict.get(new_name, "")
        with cols[1]:
            st.text_input("Craft", value=new_craft, disabled=True, key=f"new_craft{add_suffix}")
        with cols[2]:
            new_st = st.number_input("ST", min_value=0.0, step=0.5, key=f"new_st{add_suffix}")
        with cols[3]:
            new_one5 = st.number_input("x1.5 OT", min_value=0.0, step=0.5, key=f"new_one5{add_suffix}")
        with cols[4]:
            new_dt = st.number_input("x2 OT", min_value=0.0, step=0.5, key=f"new_dt{add_suffix}")
        with cols[5]:
            st.write("")
        if st.button("Add This Employee", key=f"add_employee{add_suffix}"):
            st.session_state.employees.append({
                "name": new_name,
                "craft": new_craft,
                "st": new_st,
                "one5": new_one5,
                "dt": new_dt
            })
            st.session_state.add_counter += 1
            st.rerun()

    work_notes = st.text_area("WORK PERFORMED/NOTES", height=200)

    # Equipment - checkboxes instead of numbers
    st.subheader("EQUIPMENT USED TODAY")
    col1, col2, col3 = st.columns(3)

    with col1:
        svc_truck = st.checkbox("SERVICE TRUCK/VAN")
        frm_truck = st.checkbox("FOREMAN TRUCK")
        weld_mach = st.checkbox("WELDING MACHINE")
        vac_pump = st.checkbox("VACUUM PUMP")
        gas_meter = st.checkbox("4 GAS METER")
        torch = st.checkbox("TORCH SET UP")
        orbital = st.checkbox("ORBITAL WELDER")

    with col2:
        pipe_mach = st.checkbox("PIPE MACHINE")
        pro_press = st.checkbox("PRO PRESS GUN")
        b_tank = st.checkbox("B-TANK")
        hot_tap = st.checkbox("HOT TAP MACHINE")
        plasma = st.checkbox("PLASMA CUTTER")
        hydro = st.checkbox("HYDRO PUMP")
        scissor = st.checkbox("MARTIN SCISSOR LIFT")

    with col3:
        nitro = st.checkbox("NITROGEN")
        argon_gas = st.checkbox("ARGON")
        rent1 = st.checkbox("RENTAL 1")
        rent1_type = st.text_input("Type", key="rent1_type") if rent1 else ""
        rent2 = st.checkbox("RENTAL 2")
        rent2_type = st.text_input("Type", key="rent2_type") if rent2 else ""
        rent3 = st.checkbox("RENTAL 3")
        rent3_type = st.text_input("Type", key="rent3_type") if rent3 else ""
        oth1 = st.checkbox("OTHER 1")
        oth1_type = st.text_input("Type", key="oth1_type") if oth1 else ""
        oth2 = st.checkbox("OTHER 2")
        oth2_type = st.text_input("Type", key="oth2_type") if oth2 else ""

    # Generate button
    if st.button("Generate Foreman's Report", type="primary"):
        try:
            # 1. Fill the Excel template (your existing filling code)
# ─── Fix for CoInitialize error ───
            
            wb = openpyxl.load_workbook("BlankForemanReport.xlsx")
            sheet = wb["FormansReport"]

            sheet["F3"] = day_name
            sheet["F4"] = report_date.strftime("%m/%d/%Y")
            sheet["F5"] = job_name
            sheet["F6"] = job_number
            sheet["F7"] = description
            # Fill the correct state cell based on selection
            if work_state == "Indiana":
                sheet["J5"] = "X"
                sheet["J4"] = ""  # clear Illinois if it was there
            elif work_state == "Illinois":
                sheet["J4"] = "X"
                sheet["J5"] = ""  # clear Indiana
            start_row = 10
            for idx, emp in enumerate(st.session_state.employees):
                row = start_row + idx
                sheet.cell(row=row, column=1).value = emp["name"]
                sheet.cell(row=row, column=8).value = emp["craft"]
                sheet.cell(row=row, column=9).value = emp["st"]
                sheet.cell(row=row, column=10).value = emp["one5"]
                sheet.cell(row=row, column=11).value = emp["dt"]

            sheet["A24"] = work_notes
            sheet["A24"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
            sheet.row_dimensions[24].height = 140

            def mark(cell, used):
                sheet[cell] = "X" if used else ""

            mark("A39", svc_truck)
            mark("A40", frm_truck)
            mark("A41", weld_mach)
            mark("A42", vac_pump)
            mark("A43", gas_meter)
            mark("A44", torch)
            mark("A45", orbital)
            mark("D39", pipe_mach)
            mark("D40", pro_press)
            mark("D41", b_tank)
            mark("D42", hot_tap)
            mark("D43", plasma)
            mark("D44", hydro)
            mark("D45", scissor)
            mark("G39", nitro)
            mark("G40", argon_gas)

            sheet["G41"] = rent1_type if rent1 else ""
            sheet["G42"] = rent2_type if rent2 else ""
            sheet["G43"] = rent3_type if rent3 else ""
            sheet["G44"] = oth1_type if oth1 else ""
            sheet["G45"] = oth2_type if oth2 else ""


            # 6. Offer PDF download (with your preferred filename style)
            job_name_clean = (job_name or "NoJobName").strip().replace(" ", "_").replace("/", "-")[:30]
            job_num_clean = (job_number or "NoJobNum").strip().replace(" ", "_").replace("/", "-")[:15]
            filename = f"{job_name_clean}_{job_num_clean}_{report_date.strftime('%Y-%m-%d')}.pdf"
            # Basic PDF using weasyprint (we'll make it look better next)
            from weasyprint import HTML
            html = f"""
            <html>
            <head><style>body {{ font-family: Arial; }}</style></head>
            <body>
            <h1>Foreman's Daily Report</h1>
            <p><b>State:</b> {work_state}</p>
            <p><b>Date:</b> {report_date.strftime('%m/%d/%Y')} ({day_name})</p>
            <p><b>Job:</b> {job_name} - {job_number}</p>
            <p><b>Description:</b> {description}</p>
            <h2>Employees</h2>
            <ul>
            """
            for emp in st.session_state.employees:
                html += f"<li>{emp['name']} ({emp['craft']}): ST {emp['st']}, 1.5 {emp['one5']}, DT {emp['dt']}</li>"
            html += f"""
            </ul>
            <h2>Notes:</h2>
            <p>{work_notes.replace('\n', '<br>')}</p>
            <h2>Equipment Used:</h2>
            <p>Service Truck: {'X' if svc_truck else ''}</p>
            <!-- add more equipment lines as needed -->
            </body>
            </html>
            """
            pdf_bytes = HTML(string=html).write_pdf()

            filename = f"{job_name or 'NoJob'}_{job_number or 'NoNum'}_{report_date.strftime('%Y-%m-%d')}.pdf"

            st.download_button(
                label="📥 Download PDF Report",
                data=pdf_bytes,
                file_name=filename,
                mime="application/pdf",
                key="pdf_dl"
            )
            st.download_button(
                label="📥 Download PDF Report",
                data=pdf_data,
                file_name=filename,
                mime="application/pdf",
                key="pdf_download"
            )

            st.success("PDF created successfully! Click above to download.")

# ─── WEEKLY TIMESHEET ───────────────────────────────────────────────────────
with tab2:
    st.header("Weekly Per Employee Timesheet")
    st.info("We'll build this after the Foreman's Report is done.")

st.markdown("---")
st.caption("Make sure EmployeeNames.xlsx and BlankForemanReport.xlsx are in the same folder.")

